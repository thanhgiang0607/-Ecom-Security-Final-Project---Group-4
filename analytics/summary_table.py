import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_TRUOC = "../results/bao_cao_dinh_luong_loi.csv" 
FILE_SAU = "../results/bao_cao_dinh_luong_sau_va_loi.csv"
OUTPUT_MARKDOWN = "bang_thong_ke_truoc_sau.md"
OUTPUT_EXCEL = "bang_thong_ke_truoc_sau.xlsx"

def xu_ly_du_lieu_giai_doan(file_path, ten_giai_doan):
    if not os.path.exists(file_path):
        print(f"⚠ Cảnh báo: Không tìm thấy file dữ liệu: {file_path}")
        return None
    df = pd.read_csv(file_path)
    summary = df.groupby(['Mã KB', 'Loại Tấn Công', 'HTTP Status', 'Kết Quả Thực Nghiệm']).agg(
        So_Luong=('Payload', 'count'),
        Tres_TB=('Thời Gian (s)', 'mean')
    ).reset_index()
    summary['Tres_TB'] = summary['Tres_TB'].round(4)
    summary.insert(0, 'Giai Đoạn', ten_giai_doan)
    return summary

df_truoc = xu_ly_du_lieu_giai_doan(FILE_TRUOC, "Trước")
df_sau = xu_ly_du_lieu_giai_doan(FILE_SAU, "Sau")

if df_truoc is None or df_sau is None:
    print("❌ Lỗi: Thiếu một trong hai file dữ liệu thô để đối chứng!")
    exit()

df_flat = pd.concat([df_truoc, df_sau], ignore_index=True)
df_matrix = df_flat.pivot_table(
    index=['Mã KB', 'Loại Tấn Công', 'HTTP Status', 'Kết Quả Thực Nghiệm'],
    columns=['Giai Đoạn'],
    values=['So_Luong', 'Tres_TB']
).fillna(0)

df_matrix.columns = [f"{col[0]}_{col[1]}" for col in df_matrix.columns]
df_matrix = df_matrix.reset_index()

def tinh_ti_le_giam_thieu(row):
    status = str(row['HTTP Status'])
    pre_count = row.get('So_Luong_Trước', 0)
    post_count = row.get('So_Luong_Sau', 0)
    bc_ky_thuat = str(row['Kết Quả Thực Nghiệm'])

    if status in ['200', '500', 'TIMEOUT']:
        if pre_count > 0:
            ti_le = ((pre_count - post_count) / pre_count) * 100
            return f"{ti_le:.2f}%"
        return "0.00%"
    
    elif status == '400':
        if 'Sai định dạng Captcha' in bc_ky_thuat or 'Phát hiện Payload độc hại' in bc_ky_thuat:
            return "100.00% (Vá lỗi thành công)"
        return "100.00%"
        
    elif status == '401' and 'Bị chặn bởi Bộ lọc an toàn' in bc_ky_thuat:
        return "100.00% (Màng lọc WAF RAM)"
    elif status == '201' or (status == '401' and 'Thất bại' in bc_ky_thuat):
        return "Nghiệp vụ sạch"
    else:
        return "Màng lọc bảo vệ"

df_matrix['Tỷ Lệ Giảm Thiểu Rủi Ro (R_mit)'] = df_matrix.apply(tinh_ti_le_giam_thieu, axis=1)

cols_order = [
    'Mã KB', 'Loại Tấn Công', 'HTTP Status', 'Kết Quả Thực Nghiệm',
    'So_Luong_Trước', 'So_Luong_Sau', 'Tres_TB_Trước', 'Tres_TB_Sau',
    'Tỷ Lệ Giảm Thiểu Rủi Ro (R_mit)'
]
df_matrix = df_matrix[cols_order]

academic_headers = {
    'Mã KB': 'Mã KB', 'Loại Tấn Công': 'Phân Hệ Kịch Bản', 'HTTP Status': 'Mã HTTP',
    'Kết Quả Thực Nghiệm': 'Bản Chất Kỹ Thuật / Lỗi Hệ Thống',
    'So_Luong_Trước': 'Số Lượng Trước (Reqs)', 'So_Luong_Sau': 'Số Lượng Sau (Reqs)',
    'Tres_TB_Trước': 'Thời Gian Phản Hồi TB Trước (s)', 'Tres_TB_Sau': 'Thời Gian Phản Hồi TB Sau (s)',
    'Tỷ Lệ Giảm Thiểu Rủi Ro (R_mit)': 'Tỷ Lệ Giảm Thiểu Rủi Ro (R_mit)'
}
df_matrix_view = df_matrix.rename(columns=academic_headers)

df_matrix_view.to_markdown(OUTPUT_MARKDOWN, index=False)
print(f"✔ Đã xuất kết quả tổng hợp bảo mật tại: {OUTPUT_MARKDOWN}")

try:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ma Trận So Sánh & R_mit"
    ws.views.sheetView[0].showGridLines = True

    HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    HIGHLIGHT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    font_title = Font(name="Arial", size=14, bold=True, color="1F497D")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10)
    font_bold = Font(name="Arial", size=10, bold=True)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9")
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = "MA TRẬN ĐỐI CHỨNG VÀ ĐÁNH GIÁ TỶ LỆ GIẢM THIỂU RỦI RO KIẾN TRÚC"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35

    for col_num, h_text in enumerate(df_matrix_view.columns, 1):
        cell = ws.cell(row=3, column=col_num, value=h_text)
        cell.font = font_header
        cell.fill = HEADER_FILL
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[3].height = 30

    for row_idx, row_data in enumerate(df_matrix_view.values, 4):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            if row_idx % 2 == 0: cell.fill = ZEBRA_FILL

            if col_idx in [1, 3]: cell.alignment = align_center
            elif col_idx in [2, 4]: cell.alignment = align_left
            elif col_idx in [5, 6]:
                cell.alignment = align_right
                cell.number_format = '#,##0'
            elif col_idx in [7, 8]:
                cell.alignment = align_right
                cell.number_format = '0.0000'
            elif col_idx == 9:
                cell.alignment = align_center
                if "100.00%" in str(val):
                    cell.font = font_bold
                    cell.fill = HIGHLIGHT_GREEN

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col if cell.row != 1)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(OUTPUT_EXCEL)
    print(f"✔ Đã xuất file Excel hoàn mỹ tại: {OUTPUT_EXCEL}")
except Exception as e:
    print(f"❌ Lỗi xử lý định dạng Excel: {e}")